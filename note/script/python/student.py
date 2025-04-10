# -*- coding: utf-8 -*-
from datetime import datetime, time
import openpyxl

def is_time_in_range(target_time):
    target_date = datetime(2025, 4, 10).date()
    return time(9, 0) <= target_time.time() <= time(12, 0)

def readExcel():
    my_set = set()
    
    s23da5 = {"杜子恒","陈晓天","张梓航","任静伟","李心雨","张一蕊","耿甲妍","袁鑫","付屹桐","孙熠婷","付露冉","潘德凤","李秋歌","闫玉娇","王帅","任轩鞍","王泽翾","胡家纬","王虹程","张哲","刘怡婧","张天婧","冯蕊","胡昕阳","武子昊","孟玉轩","胡紫贺","宋鑫","田浩","姜鸣一","卢静怡","徐璇","王文垚","齐文华","王俊皓","宁帅泽","韩家璇","蒲怡含","高雨寒","刘鑫越","梁紫涵","顾新旺","吴广震","梁亚坤","马光瑞","冯雅诺","王旭","付飞扬"}
    s23da6 = {"赵洛雨","赵泽天","代紫盈","陈思佳","提寿臣","王彬","万正豪","薛梦娇","孙璐瑶","李泽浩","马冰","闫静怡","葛青瑞","任爱","梁爽","杨贺斐","张梦晗","张紫玉","苏士哲","王蔓","刘兆宇","吕婧玟","孙林","张恒硕","郭天硕","张雅璇","李佩卓","李洪成","高淼","李鑫桐","王晨爽","杨紫嫣","孙一凡","李祐墨","李建伟","董佳悦","史子涵","崔佳语","张国庆","王子璇","朱怡宁","王薪凯","李颖","景淼","李博","李星安","张珂浩","高佳辉"}
    
    workbook = openpyxl.load_workbook('E://每日课堂练习（收集结果）.xlsx')
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]

    # 遍历工作表中的数据
    for row in sheet.iter_rows(min_row=2, values_only=True):  # values_only=True 返回单元格值而不是对象
        cell_time = datetime.strptime(row[0], '%Y年%m月%d日 %H:%M')
        if is_time_in_range(cell_time):
            my_set.add(row[2])
        
    print(s23da5-my_set)
    
if __name__ == "__main__":
    readExcel()
